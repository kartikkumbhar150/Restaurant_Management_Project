#!/usr/bin/env python3
import sys
import json
import os
from openpyxl import load_workbook
import csv

def parse_xlsx(file_path):
    products = []
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {h.lower(): i for i, h in enumerate(headers)}

    def get_value(row, keys):
        for key in keys:
            if key.lower() in header_map:
                val = row[header_map[key.lower()]].value
                return str(val).strip() if val is not None else ""
        return ""

    for row in sheet.iter_rows(min_row=2, values_only=False):
        product = {
            "id": parse_id(get_value(row, ["id", "ID"])),
            "name": get_value(row, ["name", "Name"]),
            "price": parse_price(get_value(row, ["price", "Price"])),
            "description": get_value(row, ["description", "Description"]),
            "category": get_value(row, ["category", "Category"]),
            "subCategory": get_value(row, ["subCategory", "SubCategory", "Subcategory"])
        }
        products.append(product)

    return products

def parse_csv(file_path):
    products = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            product = {
                "id": parse_id(row.get("id")),
                "name": row.get("name"),
                "price": parse_price(row.get("price")),
                "description": row.get("description"),
                "category": row.get("category"),
                "subCategory": row.get("subCategory")
            }
            products.append(product)
    return products

def parse_id(value):
    if not value:
        return None
    try:
        return int(value)
    except:
        return None

def parse_price(value):
    if not value:
        return 0.0
    try:
        return float(value)
    except:
        return 0.0

def parse_image(file_path):
    # your existing image logic
    return [
        {"name": "Detected Product 1", "price": 10.0, "description": "From image"},
        {"name": "Detected Product 2", "price": 25.5, "description": "From image"}
    ]

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: main.py <file_path>"}))
        sys.exit(1)

    file_path = sys.argv[1]

    if not os.path.exists(file_path):
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    try:
        if ext in [".xlsx", ".xlsm", ".xls"]:
            products = parse_xlsx(file_path)
        elif ext in [".csv"]:
            products = parse_csv(file_path)
        elif ext in [".png", ".jpg", ".jpeg"]:
            products = parse_image(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

        print(json.dumps(products, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
